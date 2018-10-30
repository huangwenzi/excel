
# 操作excel类
import os
from openpyxl import Workbook  # 新建时导入这个
from openpyxl import load_workbook  # 读取时导入这个

class Operation():

    # 读取全部文件数据到内存中，建立一个列表保存
    # 并对数据进行总的合并
    def read_data(self, data):
        
        dir_path = data.dir_path
        putout = data.putout
        begin_row = data.begin_row
        consult_key = data.consult_key
        fill_key = data.fill_key
        Retain_key = data.Retain_key
        add_key = data.add_key

        # 检查文件是否存在
        if os.path.exists(dir_path) == False:
            print("文件不存在")
            return False

        # 开始读取数据, 目前只对第一个表进行操作
        workbook = load_workbook(dir_path)
        sheet_name = workbook.get_sheet_names()[0]
        sheet = workbook[sheet_name]

        # 获取字段名
        keys_list = []
        attr_count = sheet.max_column
        for index in range(0, attr_count):
            keys_list.append(sheet.cell(row=begin_row, column=index+1).value)

        # 获取一个有效对象所占的行数
        # 获取参考键所有的列数
        consult_col = keys_list.index(consult_key)
        if consult_col == -1:
            return False
        one_consult = sheet.cell(row=begin_row+1, column=consult_col+1).value
        consult_count = 1
        for_count = 0
        while True:
            if for_count == 100:
                break
            value_1 = sheet.cell(row=begin_row + 1 + consult_count, column=consult_col+1).value
            if value_1 != one_consult:
                break
            consult_count += 1
        
        # 获取需要填0字段的所占列数
        fill_col = []
        for index in range(0, len(fill_key)):
            index_1 = keys_list.index(fill_key[index]) + 1
            if index_1 == 0:
                continue
            fill_col.append(index_1)

        # 获取需要保留字段的所占列数
        Retain_col = []
        for index in range(0, len(Retain_key)):
            index_1 = keys_list.index(Retain_key[index]) + 1
            if index_1 == 0:
                continue
            Retain_col.append(index_1)

        # 获取需要増位字段的所占列数
        add_col = []
        for index in range(0, len(add_key)):
            index_1 = keys_list.index(add_key[index]) + 1
            if index_1 == 0:
                continue
            add_col.append(index_1)

        # 开始进行复制处理
        # 创建新的excel文档
        excel = Workbook()
        excel.create_sheet('sheet1', index=0)
        sheet1 = excel['sheet1']

        # 把字段开始前的内容复制过去
        for index in range(0, begin_row):
            for index_1 in range(0,attr_count):
                value = sheet.cell(row=index+1, column=index_1+1).value
                sheet1.cell(row=index+1, column=index_1+1).value = value

        # 开始对字段后的内容正式处理
        row_count = sheet.max_row
        copy_count = begin_row+1
        index = begin_row + 1
        while index < row_count:
            # 先分析出需要插入的数据有多少
            vlaue = sheet.cell(row=index, column=consult_col+1).value
            vlaue_count = 1
            for index_1 in range(1, consult_count):
                vlaue_1 = sheet.cell(row=index + index_1, column=consult_col+1).value
                if vlaue != vlaue_1:
                    break
                vlaue_count += 1
            # 需要添加数据行数
            add_data_col = consult_count - vlaue_count

            # 先把原有的拷贝
            for index_1 in range(0, vlaue_count):
                for index_2 in range(0, attr_count):
                    value_2 = sheet.cell(row=index + index_1, column=index_2+1).value
                    sheet1.cell(row=copy_count, column=index_2+1).value = value_2
                copy_count += 1
            # 
            index += vlaue_count

            # 再把填补数据加入
            for index_1 in range(0, add_data_col):
                for index_2 in range(0, attr_count):
                    # fill_key 填充键填入
                    if index_2+1 in fill_col:
                        sheet1.cell(row=copy_count, column=index_2+1).value = 0
                    # Retain_key 保留键
                    elif index_2+1 in Retain_col:
                        value_3 = sheet1.cell(row=copy_count - 1, column=index_2+1).value
                        sheet1.cell(row=copy_count, column=index_2+1).value = value_3
                    # add_key 増位键
                    elif index_2+1 in add_col:
                        value_3 = sheet1.cell(row=copy_count - 1, column=index_2+1).value
                        sheet1.cell(row=copy_count, column=index_2+1).value = value_3 + 1
                    else :
                        pass

                copy_count += 1
            # if index >= 10:
            #     excel.save(putout)
            #     return
            

        # 最后保存数据
        # 保存excel表
        excel.save(putout)

operation = Operation()