
# 这里的作用是把toml文件按照一定格式做换成excel
import json
import toml
import os
from openpyxl import Workbook  # 新建时导入这个
from openpyxl import load_workbook  # 读取时导入这个

# 排序的条件
condition = "zzSD"

# 获取文件名并检查文件是否存在
# filename = input("输入文件地址")
# if os.path.exists(filename) == False:
#     print("文件不存在")

filename = "G:\\huangwen\\code\\excel\\新建文件夹\\poke.toml"

# 先读取toml表格里的数据, 转换成一个字典
# 加上读取的格式可以有效避免乱码
f = open(filename, encoding="utf8", errors='ignore')
poke_tab = toml.load(f)

# 根据第一个对象的key来做标准对象
# 或许可以找到一个最多元素的Excel表来做主对象
# 所以需要一开始就把全部的excel表格全部读取保存到内存中
# 做表格参数不同的兼容会消耗大量内存和cpu
obj_list = list(poke_tab.values())
sta_keys = list(obj_list[0].keys())

# 对obj_list 进行排序
obj_list.sort(key=lambda x: x[condition], reverse=False)

# 创建excel对象
excel = Workbook()
excel.create_sheet('sheet1', index=0)
sheet1 = excel['sheet1']

# 先写第一行的字段名
for index in range(0, len(sta_keys)):
    sheet1.cell(row=1, column=index+1).value = sta_keys[index]

# 把所有的数据填入
for index in range(0, len(obj_list)):
    for index_1 in range(0, len(sta_keys)):
        sheet1.cell(row=2+index, column=index_1+1).value = obj_list[index][sta_keys[index_1]]

# 保存excel表
excel.save("test.xlsx")
