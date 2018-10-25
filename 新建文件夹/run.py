# coding=utf-8

# 批量处理精灵数据

import toml
import os
from openpyxl import Workbook  # 新建时导入这个
from openpyxl import load_workbook  # 读取时导入这个

# 先读取toml表格里的数据, 转换成一个字典
# 加上读取的格式可以有效避免乱码
f = open("poke.toml", encoding = "utf8", errors = 'ignore')
poke_tab = toml.load(f)
# 获取表里的值
poke_value = list(poke_tab.values())
poke_count = len(poke_value)

# 读取要写入数据的Excel
workbook = load_workbook('./text_1.xlsx')
sheet = workbook.get_sheet_by_name('Sheet1')


# 先写第一行，精灵名字，血量，物攻，物防，特攻，特防，速度，种族值
# openpyxl的row 和  col是从1开始的，极度不友善╭(╯^╰)╮
attr = ["id", "名字", '血量', '物攻', '物防', '特攻', '特防', '速度', '种族值']
attr_count = len(attr)
for index in range(0, attr_count):
    sheet.cell(row = 1, column = index+1).value = attr[index]

# 现在开始吧数据写入
for index in range(0, poke_count):
    sheet.cell(row=index + 2, column = 1).value = poke_value[index]['id']
    sheet.cell(row=index + 2, column = 2).value = poke_value[index]['chineseName']
    sheet.cell(row=index + 2, column = 3).value = poke_value[index]['zzHP']
    sheet.cell(row=index + 2, column = 4).value = poke_value[index]['zzWG']
    sheet.cell(row=index + 2, column = 5).value = poke_value[index]['zzWF']
    sheet.cell(row=index + 2, column = 6).value = poke_value[index]['zzTG']
    sheet.cell(row=index + 2, column = 7).value = poke_value[index]['zzTF']
    sheet.cell(row=index + 2, column = 8).value = poke_value[index]['zzSD']
    # 求和种族值, 前两个不算
    sum = 0
    for index_1 in range(2, attr_count - 1):
        sum += sheet.cell(row=index + 2, column = index_1 + 1).value
    sheet.cell(row=index + 2, column = 9).value = sum

# 排序

# 保存
workbook.save('./text_2.xlsx')
