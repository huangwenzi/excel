
# 这个是执行excel功能的类

import os
from openpyxl import Workbook  # 新建时导入这个
from openpyxl import load_workbook  # 读取时导入这个
from config import config


class Excel():

    # 检查目录
    # dir_name : 路径名
    # 返回目录下的excel格式的文件名列表
    def check_dir(self, data):
        # 检查路径是否存在
        if os.path.exists(data.dir_path) == False:
            print("路径不存在")
            return False

        # 获取目录下的文件名
        file_name_list = os.listdir(data.dir_path)
        print("目录下有文件:")
        print(file_name_list)

        # 查找符合的文件
        name_list = []
        for name in file_name_list:
            if name.find(".xlsx") != -1:
                name_list.append(name)
        print("符合条件的文件有:")
        print(name_list)

        # 数据保存
        data.name_list = name_list

        return True

    # 读取全部文件数据到内存中，建立一个列表保存
    # 并对数据进行总的合并

    def read_data(self, data):
        # 检查文件列表
        if len(data.name_list) <= 0:
            print("文件列表为空")
            return False

        name_list = data.name_list

        # 开始读取数据
        sheet_list = []
        for index in range(0, len(name_list)):
            workbook = load_workbook(
                data.dir_path + "/" + name_list[index])
            sheet_name = workbook.get_sheet_names()[0]
            sheet_list.append(workbook[sheet_name])

        # 获取每个列表的属性名列表
        keys_list = []
        for index in range(0, len(name_list)):
            attr_count = sheet_list[index].max_column
            # print(attr_count)
            tmp = []
            for index_1 in range(0, attr_count):
                tmp.append(sheet_list[index].cell(
                    row=1, column=index_1+1).value)

            keys_list.append(tmp)
            # print(tmp)

        # 需要把每一行数据都转换成字典对象的形式保存才好操作
        dict_list = []
        # 这层循环是表格数的
        for index in range(0, len(name_list)):
            row_list = []
            row_count = sheet_list[index].max_row
            attr_count = len(keys_list[index])
            # 这层循环是行数的
            for index_1 in range(1, row_count):
                row_dict = {}
                # 这层循环是列数的
                for index_2 in range(0, attr_count):
                    row_dict[keys_list[index][index_2]] = sheet_list[index].cell(
                        row=index_1+1, column=index_2+1).value
                row_list.append(row_dict)
            dict_list.append(row_list)

        # 属性列表合并去重，获得组合在一起的包含全部的属性表
        tmp_attr = []
        attr_list = []
        for index in range(0, len(keys_list)):
            tmp_attr += keys_list[index]
        for index in range(0, len(tmp_attr)):
            if tmp_attr[index] not in attr_list:
                attr_list.append(tmp_attr[index])

        # print("去重前的属性表")
        # print(tmp_attr)
        # print("去重后的属性表")
        # print(attr_list)

        # 数据表之间的合并
        data_list = []
        for index in range(0, len(dict_list)):
            data_list += dict_list[index]

        # 删除主键不存在的非法数据
        key = config.primary_key
        count = len(data_list)
        tmp_index = 0
        for index in range(0, count):
            if data_list[tmp_index][key] == None:
                data_list.pop(tmp_index)
                tmp_index -= 1
            tmp_index += 1

        # 保存数据到内存
        # 数据行的字典总表
        data.data_list = data_list
        # 属性列表
        data.attr_list = attr_list

    # 排序处理
    # order_key :　根据此键排序
    # order : 升序还是降序
    def sort_data(self, data):

        # 检查是不是不做排序处理
        if data.order_key == "":
            return

        data_list = data.data_list
        attr_list = data.attr_list
        order_key = data.order_key
        order = data.order

        # 先检查排序的字段合不合法
        if order_key not in attr_list:
            print("填入的排序键不存在")
            return

        # 对不存在排序键值的进行填零处理
        add_conut = 0
        for index in range(0, len(data_list)):
            if data_list[index][order_key] == None:
                data_list[index][order_key] = 0
                add_conut += 1

        # 对数据进行排序
        # 要解决一下某些不存在这个键的问题
        data_list.sort(key=lambda x: x[order_key], reverse=order)

        # 对进行过填0处理的数据恢复为空
        if order:
            count = len(data_list)
            for index in range(0, add_conut):
                data_list[count - index - 1][order_key] = None
        else:
            for index in range(0, add_conut):
                data_list[index][order_key] = None

        # 保存修改后的数据
        data.data_list = data_list

    # 保存数据
    # path : 保存的路径
    # file_name : 文件名
    def save(self, data):

        attr_list = data.attr_list
        data_list = data.data_list
        path = data.putout
        file_name = data.file_name

        # 先检查路径是否存在
        if os.path.exists(path) == False:
            print("路径不存在")
            return

        # 创建新的excel文档
        excel = Workbook()
        excel.create_sheet('sheet1', index=0)
        sheet1 = excel['sheet1']

        # 先写入第一行
        for index in range(0, len(attr_list)):
            sheet1.cell(row=1, column=index+1).value = attr_list[index]

        # 根据去重后的属性表进行数据的输入
        # 这层循环是整合的表数
        row_count = 2   # 当前应该写入的行数
        for index in range(0, len(data_list)):
            tmp_keys = data_list[index].keys()
            tmp_row = data_list[index]
            # 这层循环是属性数
            for index_1 in range(0, len(attr_list)):
                if attr_list[index_1] not in tmp_keys:
                    continue
                sheet1.cell(row=row_count, column=index_1 +
                            1).value = tmp_row[attr_list[index_1]]
            row_count += 1

        # 保存excel表
        path = path + "/" + file_name
        excel.save(path)


# 实例化单例
excel = Excel()
